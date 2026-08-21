import io
import csv
import uuid
import datetime
import hashlib
from typing import List, Dict, Any, Optional, Tuple
from fastapi import HTTPException
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from starlette.concurrency import run_in_threadpool

from app.services.flashcard_service import (
    get_supabase,
    in_memory_decks,
    in_memory_cards,
    to_valid_uuid
)
from app.rabbitmq_consumer import publish_flashcard_created_event
from app.utils.logger import logger

# Header aliases for smart column detection across multiple languages
HEADER_ALIASES = {
    "term": [
        # Vietnamese & English
        "term", "từ vựng", "tu vung", "word", "front", "từ gốc", "tu goc", "vocabulary", "từ khóa", "tu khoa", "từ",
        # German
        "wort", "vokabel", "stichwort", "ausdruck",
        # French & Spanish
        "mot", "vocabulaire", "terme", "palabra", "termino", "vocablo",
        # Chinese (Simplified & Traditional)
        "生词", "生詞", "单词", "單詞", "词汇", "詞彙", "词", "詞",
        # Japanese
        "単語", "たんご", "言葉", "ことば", "見出し語",
        # Korean & Russian
        "단어", "어휘", "слово", "термин", "лексема"
    ],
    "definition": [
        # Vietnamese & English
        "definition", "nghĩa", "nghia", "nghĩa tiếng việt", "nghia tieng viet", "meaning", "back", "dịch nghĩa", "dich nghia", "giải nghĩa", "giai nghia", "ý nghĩa", "y nghia",
        # German
        "bedeutung", "übersetzung", "uebersetzung", "erklärung", "erklærung",
        # French & Spanish
        "sens", "signification", "traduction", "définition", "definicion", "significado", "traducción", "traduccion",
        # Chinese
        "意思", "释义", "釋義", "翻译", "翻譯", "解释", "解釋",
        # Japanese
        "意味", "いみ", "訳", "やく", "定義", "日本語訳",
        # Korean & Russian
        "뜻", "의미", "번역", "значение", "перевод", "определение"
    ],
    "phonetic": [
        # Vietnamese & English
        "phonetic", "phiên âm", "phien am", "ipa", "pronunciation", "phát âm", "phat am", "cách đọc", "cach doc",
        # German & French & Spanish
        "aussprache", "lautschrift", "prononciation", "pronunciacion", "pronunciación", "fonética", "fonetica",
        # Asian Pronunciation (Pinyin, Furigana, Romaji, Hangul Romanization)
        "pinyin", "bính âm", "拼音", "注音", "ふりがな", "furigana", "romaji", "ローマ字", "読み方", "よみかた", "발음", "로마자",
        # Russian
        "произношение", "транскрипция", "фонетика"
    ],
    "part_of_speech": [
        # Vietnamese & English
        "part_of_speech", "partofspeech", "từ loại", "tu loai", "type", "pos", "loại từ", "loai tu",
        # German & French & Spanish
        "wortart", "nature", "type_de_mot", "categoria_gramatical", "categoría gramatical", "clase_de_palabra",
        # Asian & Russian
        "词性", "詞性", "品詞", "ひんし", "품사", "часть речи", "категория"
    ],
    "example": [
        # Vietnamese & English
        "example", "ví dụ", "vi du", "example_sentence", "examplesentence", "context", "câu ví dụ", "cau vi du", "ngữ cảnh", "ngu canh", "mẫu câu", "mau cau",
        # German & French & Spanish
        "beispiel", "beispielsatz", "anwendungsbeispiel", "exemple", "phrase_exemple", "ejemplo", "frase_de_ejemplo", "contexto",
        # Asian & Russian
        "例句", "例", "例文", "れいぶん", "예문", "문장", "пример", "контекст", "предложение"
    ]
}

def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()

def _match_header(header_name: str) -> Optional[str]:
    cleaned = header_name.strip().lower()
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in cleaned or cleaned in alias:
                return field
    return None

def parse_excel_or_csv(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Phân tích file Excel (.xlsx, .xls) hoặc CSV:
    - Tự động nhận diện header tiếng Anh/tiếng Việt
    - Lọc bỏ dòng trống, trim khoảng trắng
    - Deduplicate các từ vựng trùng nhau trong cùng file
    - Trả về danh sách cards chuẩn hóa
    """
    fname = filename.lower()
    raw_rows: List[List[str]] = []

    if fname.endswith(".csv"):
        # Thử các bộ mã hóa phổ biến
        text_content = ""
        for encoding in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
            try:
                text_content = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if not text_content:
            raise HTTPException(status_code=400, detail="Không thể đọc nội dung file CSV. Vui lòng lưu file ở định dạng UTF-8.")

        reader = csv.reader(io.StringIO(text_content))
        for row in reader:
            if any(cell.strip() for cell in row):
                raw_rows.append([cell.strip() for cell in row])

    elif fname.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            if sheet is None:
                raise HTTPException(status_code=400, detail="File Excel không chứa trang tính (sheet) nào.")

            for row in sheet.iter_rows(values_only=True):
                str_row = [_clean_str(cell) for cell in row]
                if any(str_row):
                    raw_rows.append(str_row)
        except Exception as e:
            logger.error(f"Error parsing Excel file {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Không thể mở file Excel. Lỗi: {str(e)}")
    else:
        raise HTTPException(
            status_code=400,
            detail="Định dạng file không được hỗ trợ. Vui lòng tải lên file .xlsx, .xls hoặc .csv."
        )

    if not raw_rows:
        raise HTTPException(status_code=400, detail="File tải lên không có dữ liệu.")

    # Xác định Header Mapping
    first_row = raw_rows[0]
    header_mapping: Dict[int, str] = {}
    has_header = False

    for idx, cell in enumerate(first_row):
        matched = _match_header(cell)
        if matched:
            header_mapping[idx] = matched
            has_header = True

    # Kiểm tra xem có nhận diện được ít nhất 1 cột bắt buộc không
    detected_fields = set(header_mapping.values())
    start_index = 1 if has_header else 0

    if not has_header or "term" not in detected_fields:
        # Fallback sang Positional Mapping mặc định
        header_mapping = {
            0: "term",
            1: "definition",
            2: "phonetic",
            3: "part_of_speech",
            4: "example"
        }
        # Nếu dòng đầu là text header không khớp, bỏ qua dòng đầu
        if any(h in first_row[0].lower() for h in ["term", "từ", "word", "stt"]):
            start_index = 1
        else:
            start_index = 0

    parsed_cards: List[Dict[str, Any]] = []
    seen_terms = set()

    for row_idx in range(start_index, len(raw_rows)):
        row = raw_rows[row_idx]
        card_obj: Dict[str, Any] = {
            "term": "",
            "definition": "",
            "phonetic": "",
            "part_of_speech": "noun",
            "example": ""
        }

        for col_idx, field_name in header_mapping.items():
            if col_idx < len(row):
                val = _clean_str(row[col_idx])
                if val:
                    card_obj[field_name] = val

        term = card_obj.get("term", "").strip()
        definition = card_obj.get("definition", "").strip()

        # Bỏ qua nếu thiếu từ gốc hoặc nghĩa
        if not term or not definition:
            continue

        # Bỏ qua nếu từ gốc là header lặp lại
        if term.lower() in ["term", "từ vựng", "word", "từ gốc"]:
            continue

        # Chuẩn hóa từ loại đa ngôn ngữ (Vietnamese, English, German, French, Spanish, Chinese, Japanese, Korean, Russian)
        pos = card_obj.get("part_of_speech", "").lower()
        if any(k in pos for k in ["danh", "noun", "nomen", "nom", "sustantivo", "名词", "名詞", "명사", "сущ", "n."]):
            card_obj["part_of_speech"] = "noun"
        elif any(k in pos for k in ["động", "dong", "verb", "verbe", "verbo", "动词", "動詞", "동사", "глагол", "v."]):
            card_obj["part_of_speech"] = "verb"
        elif any(k in pos for k in ["tính", "tinh", "adj", "adjektiv", "adjectif", "adjetivo", "形容", "형용사", "прил", "a."]):
            card_obj["part_of_speech"] = "adjective"
        elif any(k in pos for k in ["cụm", "cum", "phrase", "phr", "phrasen", "locution", "frase", "短语", "句", "숙어", "фраза"]):
            card_obj["part_of_speech"] = "phrase"
        else:
            card_obj["part_of_speech"] = "phrase" if " " in term else "noun"

        # Deduplicate trong cùng file
        term_key = term.lower()
        if term_key in seen_terms:
            continue
        seen_terms.add(term_key)

        parsed_cards.append(card_obj)

        if len(parsed_cards) >= 2000:
            break

    if not parsed_cards:
        raise HTTPException(
            status_code=400,
            detail="Không tìm thấy thẻ từ vựng hợp lệ nào trong file. Đảm bảo file có ít nhất cột 'Term' (Từ vựng) và 'Definition' (Định nghĩa)."
        )

    return parsed_cards

def generate_excel_template() -> bytes:
    """
    Tạo file Excel mẫu chuẩn (.xlsx) hỗ trợ đa ngôn ngữ có styling và các dòng ví dụ minh họa.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flashcard_Template"

    # Styling
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Blue-600
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = [
        "Term (Từ vựng ngoại ngữ)*",
        "Definition (Nghĩa tiếng Việt)*",
        "Phonetic (Phiên âm / IPA / Pinyin / Furigana)",
        "PartOfSpeech (Từ loại: noun/verb/adj/phrase)",
        "Example (Câu ví dụ minh họa)"
    ]

    ws.append(headers)

    # Multilingual sample rows (Tiếng Anh, Tiếng Đức, Tiếng Pháp, Tiếng Trung, Tiếng Nhật)
    sample_data = [
        ["Artificial Intelligence", "Trí tuệ nhân tạo", "/ˌɑː.tɪˈfɪʃ.əl ɪnˈtel.ɪ.dʒəns/", "noun", "Artificial Intelligence is transforming modern education."],
        ["die Universität", "Trường đại học", "/univɛʁziˈtɛːt/", "noun", "Ich studiere an der Universität."],
        ["bonjour", "Xin chào / Chúc một ngày tốt lành", "/bɔ̃.ʒuʁ/", "phrase", "Bonjour, comment allez-vous?"],
        ["你好", "Xin chào", "nǐ hǎo", "phrase", "你好！很高兴认识你。"],
        ["勉強する", "Học tập, nghiên cứu", "べんきょうする / benkyousuru", "verb", "毎日日本語を勉強します。"]
    ]

    for row in sample_data:
        ws.append(row)

    # Style Header
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Style Data rows
    for row in ws.iter_rows(min_row=2, max_row=len(sample_data) + 1):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 18)

    ws.row_dimensions[1].height = 28

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

async def bulk_insert_cards(
    deck_id: str,
    user_id: str,
    cards_data: List[Dict[str, Any]],
    lang_code: str = "en"
) -> Dict[str, Any]:
    """
    Thực hiện Bulk Insert vào Database (Supabase PostgreSQL / In-memory):
    - Kiểm tra và tự động đồng bộ Deck (IDOR protection & Local Storage Fallback)
    - Deduplicate với các từ đã có trong bộ thẻ
    - Batch chunk insert 200 bản ghi/lần
    - Bắn event RabbitMQ sinh Audio ngầm
    """
    supabase = get_supabase()
    deck_owner = None
    deck_lang = lang_code
    deck_title = "Sổ từ vựng"

    # 1. Tìm thông tin Deck (In-memory & Supabase)
    if deck_id in in_memory_decks:
        mem_deck = in_memory_decks[deck_id]
        deck_owner = str(mem_deck.get("user_id", ""))
        deck_lang = str(mem_deck.get("lang_code") or lang_code)
        deck_title = str(mem_deck.get("title") or deck_title)

    if supabase:
        try:
            valid_deck_uuid = to_valid_uuid(deck_id) or deck_id
            deck_res = await run_in_threadpool(
                lambda: supabase.table("flashcard_decks").select("id, user_id, lang_code, title").eq("id", valid_deck_uuid).execute()
            )
            if deck_res.data and len(deck_res.data) > 0:
                sb_d = deck_res.data[0]
                deck_owner = str(sb_d.get("user_id", "") or deck_owner or "")
                deck_lang = str(sb_d.get("lang_code") or deck_lang or lang_code)
                deck_title = str(sb_d.get("title") or deck_title)
            else:
                # Sổ thẻ chưa có trên Supabase (được tạo ở Client / LocalStorage)
                # Tự động tạo Deck trên Supabase để thỏa mãn Foreign Key constraint
                valid_uid = to_valid_uuid(user_id)
                auto_deck_payload = {
                    "id": valid_deck_uuid,
                    "title": deck_title,
                    "description": f"Sổ từ vựng ({deck_lang.upper()})",
                    "lang_code": deck_lang,
                    "icon_flag": "🌐"
                }
                if valid_uid:
                    auto_deck_payload["user_id"] = valid_uid

                try:
                    await run_in_threadpool(
                        lambda: supabase.table("flashcard_decks").insert(auto_deck_payload).execute()
                    )
                    logger.info(f"Auto-created missing deck {valid_deck_uuid} on Supabase during Excel import")
                except Exception as ex_deck:
                    logger.warning(f"Could not auto-insert missing deck on Supabase: {ex_deck}")
        except Exception as e:
            logger.warning(f"Error checking deck ownership in Supabase: {e}")

    # Đảm bảo in-memory deck luôn tồn tại
    if deck_id not in in_memory_decks:
        in_memory_decks[deck_id] = {
            "id": deck_id,
            "user_id": user_id,
            "title": deck_title,
            "description": f"Sổ từ vựng ({deck_lang.upper()})",
            "lang_code": deck_lang,
            "icon_flag": "🌐",
            "cards_count": 0
        }

    # IDOR Check: Chỉ chặn nếu deck thuộc sở hữu của người dùng khác
    if deck_owner and deck_owner != "anonymous" and deck_owner != str(user_id):
        valid_u = to_valid_uuid(user_id)
        if valid_u and str(valid_u) != deck_owner:
            raise HTTPException(
                status_code=403,
                detail="Bạn không có quyền thêm thẻ vào sổ thẻ này (IDOR Protection)."
            )

    # 2. Lấy danh sách từ đã có để tránh trùng lặp
    existing_terms = set()
    if supabase:
        try:
            valid_deck_uuid = to_valid_uuid(deck_id) or deck_id
            existing_res = await run_in_threadpool(
                lambda: supabase.table("flashcards").select("term").eq("deck_id", valid_deck_uuid).execute()
            )
            if existing_res.data:
                for item in existing_res.data:
                    if item.get("term"):
                        existing_terms.add(item["term"].strip().lower())
        except Exception as e:
            logger.warning(f"Error fetching existing terms in deck: {e}")

    for c in in_memory_cards.values():
        if str(c.get("deck_id")) == str(deck_id) and c.get("term"):
            existing_terms.add(c["term"].strip().lower())

    # 3. Chuẩn bị payloads cho batch insert
    now_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
    clean_lang = (deck_lang or "en").strip().lower()[:2]

    cards_to_insert: List[Dict[str, Any]] = []
    skipped_duplicates = 0

    for item in cards_data:
        term = item.get("term", "").strip()
        definition = item.get("definition", "").strip()
        if not term or not definition:
            continue

        if term.lower() in existing_terms:
            skipped_duplicates += 1
            continue

        existing_terms.add(term.lower())
        card_id = str(uuid.uuid4())

        # CAS Deduplicated Persistent Audio URL
        term_hash = hashlib.md5(f"{clean_lang}_{term.lower()}".encode('utf-8')).hexdigest()
        audio_url = f"/api/v1/translate/audio/terms/{clean_lang}_{term_hash}.mp3"

        card_dict = {
            "id": card_id,
            "deck_id": deck_id,
            "user_id": user_id,
            "term": term,
            "definition": definition,
            "phonetic": item.get("phonetic") or None,
            "audio_url": audio_url,
            "example": item.get("example") or None,
            "example_sentence": item.get("example") or None,
            "part_of_speech": item.get("part_of_speech") or "noun",
            "lang_code": deck_lang,
            # FSRS Default State
            "state": 0,
            "reps": 0,
            "lapses": 0,
            "stability": 0.0,
            "difficulty": 0.0,
            "elapsed_days": 0,
            "scheduled_days": 0,
            "last_review": None,
            "due": now_iso
        }
        cards_to_insert.append(card_dict)

    if not cards_to_insert:
        return {
            "total_in_file": len(cards_data),
            "inserted": 0,
            "skipped_duplicates": skipped_duplicates,
            "cards": []
        }

    # 4. Batch Insert (Chunk 200 bản ghi/lần)
    CHUNK_SIZE = 200
    if supabase:
        try:
            for i in range(0, len(cards_to_insert), CHUNK_SIZE):
                chunk = cards_to_insert[i:i + CHUNK_SIZE]
                sb_chunk = []
                for c in chunk:
                    valid_uid = to_valid_uuid(c.get("user_id"))
                    valid_did = to_valid_uuid(c.get("deck_id")) or c.get("deck_id")
                    sb_card = {
                        "id": c["id"],
                        "deck_id": valid_did,
                        "term": c["term"],
                        "definition": c["definition"],
                        "phonetic": c["phonetic"],
                        "audio_url": c["audio_url"],
                        "example": c["example"],
                        "part_of_speech": c["part_of_speech"],
                        "lang_code": c["lang_code"],
                        "state": 0,
                        "reps": 0,
                        "lapses": 0,
                        "stability": 0.0,
                        "difficulty": 0.0,
                        "elapsed_days": 0,
                        "scheduled_days": 0,
                        "due": now_iso
                    }
                    if valid_uid:
                        sb_card["user_id"] = valid_uid
                    sb_chunk.append(sb_card)

                await run_in_threadpool(
                    lambda: supabase.table("flashcards").insert(sb_chunk).execute()
                )
        except Exception as e:
            logger.error(f"Failed to batch insert cards to Supabase: {e}")

    # Đồng bộ vào in-memory store
    for c in cards_to_insert:
        in_memory_cards[c["id"]] = c

    if deck_id in in_memory_decks:
        in_memory_decks[deck_id]["cards_count"] = len([c for c in in_memory_cards.values() if str(c.get("deck_id")) == str(deck_id)])

    # 5. Kích hoạt sự kiện RabbitMQ để sinh âm thanh ngầm
    for c in cards_to_insert:
        try:
            await publish_flashcard_created_event(
                card_id=c["id"],
                term=c["term"],
                lang_code=c["lang_code"],
                user_id=user_id,
                phonetic=c.get("phonetic")
            )
        except Exception as e:
            logger.warning(f"Could not publish event for imported card {c['id']}: {e}")

    logger.info(f"Bulk imported {len(cards_to_insert)} cards into deck {deck_id} (Skipped {skipped_duplicates} duplicates)")

    return {
        "total_in_file": len(cards_data),
        "inserted": len(cards_to_insert),
        "skipped_duplicates": skipped_duplicates,
        "cards": cards_to_insert
    }
